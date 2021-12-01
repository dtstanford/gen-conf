#!/usr/bin/env python3
 
import argparse
from datetime import date
import ipaddress
import os
import re
import sys
import openpyxl

# Set some default global vars for use in argparse
WS_NAME = 'ACL REQUEST FORM'
VL_WS_NAME = 'Data'
START_CELL='A3'
VL_RETURN_COLS = [2, 3, 4]

# Set some global vars
CWD = os.getcwd()

def print_notice():
    notice =  '                                   *NOTICE*                                   \n'
    notice += 'This script is intended to simplify the ACL change request process. However, \n'
    notice += 'it is the user\'s responsibility to validate the configuration prior to final \n'
    notice += 'implementation. USE WITH CAUTION.'

    print()
    print(notice)
    print()
    response = input('Continue? [Y/n]: ')
    print()

    if response == '' or response.lower() == 'y':
        return
    else:
        sys.exit(0)

def parse_args():
    parser = argparse.ArgumentParser(description='A command-line tool for Static1 processing of a \
        particular client\'s ACL request forms.',)

    parser.add_argument('file', type=str, help='The .xlsx file to be processed.')
    parser.add_argument('acl_name', type=str, help='The name of the firewall ACL this will be \
                        applied to.')
    parser.add_argument('--sheet', type=str, help='The worksheet containing the ACL request form. \
                        (default: \'ACL REQUEST FORM\')', default=WS_NAME)
    parser.add_argument('--vl-sheet', type=str, help='The worksheet containing the VLOOKUP \
                        referenced data. (default: \'Data\')', default=VL_WS_NAME)
    parser.add_argument('--outfile', type=str, help='Write output to a file in addition to the \
                        screen.')
    parser.add_argument('--start-cell', type=str, help='The upper-leftmost worksheet cell to \
                        begin processing ACL rules from (e.g., \'B4\'). (default: A3)',
                        default=START_CELL)

    args = parser.parse_args(['eqfwchange-1.xlsx', 'acl_in', '--outfile', 'output'])

    return args

def validate_start_cell(start_cell):
    def error_and_exit():
        print()
        print('The supplied start-cell argument (\'{start_cell}\') is not supported.'
            .format(start_cell=start_cell))
        print()
        print('    ACCEPTABLE START-CELL RANGE: A1 -> A99')
        print()

        sys.exit(1)

    if start_cell[0].isalpha() and start_cell[1:].isnumeric():
        format_correct = True
    else:
        format_correct = False

    if 2 <= len(start_cell) <= 3:
        length_correct = True
    else:
        length_correct = False

    if format_correct and length_correct:
        return start_cell
    else:
        error_and_exit()

def validate_ip_protos(ip_protos):
    validated_ip_protos = []

    if 0 < len(ip_protos) < 3:
        for ip_proto in ip_protos:
            if ip_proto == 'tcp':
                validated_ip_protos.append(ip_proto)
            elif ip_proto == 'udp':
                validated_ip_protos.append(ip_proto)
            else:
                print()
                print('There was an error attempting to process the IP protocol: {ip_proto}'\
                    .format(ip_proto=ip_proto))
                print()

                sys.exit(1)

    return validated_ip_protos

def validate_ip_protos_ports(ip_protos_ports):
    validated_ip_protos_ports = []

    for ip_protos_port in ip_protos_ports:
        if ip_protos_port.isnumeric():
            validated_ip_protos_ports.append(ip_protos_port)
            continue
        ip_protos_port_list = ip_protos_port.split('-')
        if len(ip_protos_port_list) == 2:
            ip_protos_port_list[:] = [ip_protos_port.strip() for ip_protos_port in \
                ip_protos_port_list]
            if ip_protos_port_list[0].isnumeric() and ip_protos_port_list[1].isnumeric():
                if int(ip_protos_port_list[0]) < int(ip_protos_port_list[1]):
                    validated_ip_protos_ports.append(ip_protos_port)
                    continue
        else:
            print()
            print('There was an error attempting to process the TCP and/or UDP port(s): {ports}'\
                .format(ports=ip_protos_port))
            print()

            sys.exit(1)

    return validated_ip_protos_ports

def try_load_workbook(file):
    try:
        wb = openpyxl.load_workbook(filename=file)
    except openpyxl.utils.exceptions.InvalidFileException:
        print()
        print('The input file you supplied (\'{file}\') doesn\'t have a valid extension:'
            .format(file=file))
        print()
        print('    VALID EXTENSIONS: .xlsx, .xlsm, .xltx, .xltm')
        print()

        sys.exit(1)
    except FileNotFoundError:
        print()
        print('Unable to locate the supplied file in the current directory.')
        print()
        print('    Attempted to open: {cwd}/{file}'.format(cwd=CWD, file=file))
        print()

        sys.exit(1)

    return wb

def try_load_worksheet(wb, ws_name, wb_name):
    try:
        ws = wb[ws_name]
    except KeyError:
        print()
        print('The worksheet \'{worksheet}\' doesn\'t exist in the input file \'{file}\'.'\
            .format(worksheet=ws_name, file=wb_name))
        print()

        sys.exit(1)

    return ws

def convert_alpha_to_num(letter):
    return ord(letter.lower()) - 96

def parse_to_list(string):
    split_string = re.split(r',|\n', string)
    split_string[:] = [elem.strip() for elem in split_string]

    return split_string

def generate_net_desc(net_desc):
    net_desc = net_desc[0:53]
    net_desc = net_desc.replace(' ', '-')
    todays_date = date.today().strftime('%m-%d-%Y')
    net_desc = net_desc + '_' + todays_date

    return net_desc

def try_generate_net(net):
    try:
        net =ipaddress.IPv4Network(net)
    except:
        print()
        print('There was an error attempting to process the host/network: {net}'.format(net=net))
        print()

        sys.exit(1)
    
    return net

def vl_mapper(ws, index, return_cols):
    for row in ws.iter_rows(values_only=True):
        if index in row:
            mapped_values = []
            for col in return_cols:
                mapped_values.append(str(row[col - 1]))
            break

    return mapped_values

def generate_objgrp_config(nets, nets_desc):
    host_list = []
    net_list = []
    
    for net in nets:
        net = try_generate_net(net)
        if net.num_addresses == 1:
            host = str(net.hosts()[0])
            host_list.append(host)
        else:
            net = net.with_netmask.replace('/', ' ')
            net_list.append(net)
    
    nets_desc = generate_net_desc(nets_desc)
    objgrp_config = 'object-group network ' + nets_desc + '\n'
    
    for host in host_list:
        objgrp_config += '  network-object host ' + host + '\n'
    
    for net in net_list:
        objgrp_config += '  network-object ' + net + '\n'

    return objgrp_config

def generate_acl_config(acl, dest_nets, ip_protos, ip_protos_ports, src_net_objgrp_name):
    host_list = []
    net_list = []
    
    for net in dest_nets:
        net = try_generate_net(net)
        if net.num_addresses == 1:
            host = str(net.hosts()[0])
            host_list.append(host)
        else:
            net = net.with_netmask.replace('/', ' ')
            net_list.append(net)
    
    ip_protos = parse_to_list(ip_protos)
    ip_protos = validate_ip_protos(ip_protos)
    ip_protos_ports = parse_to_list(ip_protos_ports)
    ip_protos_ports = validate_ip_protos_ports(ip_protos_ports)
    objgrp = src_net_objgrp_name

    acl_line_prefix = 'access-list {acl} line 1 extended permit '.format(acl=acl)
    acl_config = ''

    for ip_protos_port in ip_protos_ports:
        if '-' in ip_protos_port:
            ip_protos_port = ip_protos_port.replace('-', ' ')
            port_arg = 'range'
        else:
            port_arg = 'eq'

        for ip_proto in ip_protos:
            if host_list:
                for host in host_list:
                    acl_config += acl_line_prefix \
                                + '{proto} object-group {objgrp} host {host} {port_arg} {port}\n'\
                                    .format(proto=ip_proto, objgrp=objgrp, host=host,
                                        port_arg=port_arg, port=ip_protos_port)
            if net_list:
                for net in net_list:
                    acl_config += acl_line_prefix \
                                + '{proto} object-group {objgrp} {net} {port_arg} {port}\n'\
                                    .format(proto=ip_proto, objgrp=objgrp, net=net, 
                                        port_arg=port_arg, port=ip_protos_port)
    
    return acl_config


def main():
    args = parse_args()
    print_notice()
    file = args.file
    acl = args.acl_name
    ws_name = args.sheet
    vl_ws_name = args.vl_sheet
    outfile = args.outfile
    start_cell = validate_start_cell(args.start_cell)
    start_col = convert_alpha_to_num(start_cell[0])
    start_row = int(start_cell[1:])
    return_cols = VL_RETURN_COLS

    wb = try_load_workbook(file)
    ws = try_load_worksheet(wb, ws_name, file)
    vl_ws = try_load_worksheet(wb, vl_ws_name, file)

    src_nets_col = start_col
    src_net_desc_col = start_col + 1
    dest_nets_name_col = start_col + 3

    objgrp_config_chunks = set()
    acl_config_chunks = set()

    for row in ws.iter_rows(min_col=start_col, min_row=start_row, values_only=True):
        src_nets_cell = row[src_nets_col - 1]
        src_net_desc_cell = row[src_net_desc_col - 1]
        dest_nets_name_cell = row[dest_nets_name_col - 1]

        dest_nets_cell, ip_protos_cell, ip_protos_ports_cell = vl_mapper(vl_ws, 
            dest_nets_name_cell, return_cols)
        
        src_nets_cell = parse_to_list(src_nets_cell)
        dest_nets_cell = parse_to_list(dest_nets_cell)
        src_nets_objgrp_config = generate_objgrp_config(src_nets_cell, src_net_desc_cell)
        src_net_objgrp_name = generate_net_desc(src_net_desc_cell)

        acl_config = generate_acl_config(acl, dest_nets_cell, ip_protos_cell, ip_protos_ports_cell,
            src_net_objgrp_name)

        objgrp_config_chunks.add(src_nets_objgrp_config)
        acl_config_chunks.add(acl_config)

    for objgrp_config in objgrp_config_chunks:
        print(objgrp_config)

    for acl_config in acl_config_chunks:
        print(acl_config)
    
    if outfile:
        if not outfile.endswith('.txt'):
            outfile = outfile + '.txt'
        
        with open(outfile, 'w') as file:
            for objgrp_config in objgrp_config_chunks:
                print(objgrp_config, file=file)

            for acl_config in acl_config_chunks:
                print(acl_config, file=file)

if __name__ == '__main__':
    main()